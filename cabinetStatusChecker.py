import os
import urllib.request
import zipfile

import pathlib
import glob
import time
from openpyxl import load_workbook
from datetime import date

debug = 1

print("Sto scaricando il file!\n")

#call to TIM servers
downloadUrl = 'https://www.wholesale.telecomitalia.com/sitepub/SFTP/59_Coperture_Bitstream_NGA_e_VULA/Copertura%20attiva%20e%20pianificata%20FTTCab.zip '
urllib.request.urlretrieve(downloadUrl, 'D:\Documenti\Davide\GDriveMagistrale\TIM\\2021\down.zip')

if debug == 1: print("File scaricato! Inizio estrazione!\n")

#unzip downloaded file operation
with zipfile.ZipFile('D:\Documenti\Davide\GDriveMagistrale\TIM\\2021\down.zip', 'r') as zip_ref:
    zip_ref.extractall('D:\Documenti\Davide\GDriveMagistrale\TIM\\2021')
    time.sleep(2)

if debug == 1: print("File estratto! Cancello lo zip\n")

#delete unusefull zip file
if os.path.exists('D:\Documenti\Davide\GDriveMagistrale\TIM\\2021\down.zip'):
    os.remove('D:\Documenti\Davide\GDriveMagistrale\TIM\\2021\down.zip')
    if debug == 1: print("File cancellato\n")
else:
    if debug == 1: print("Il fxile non esiste\n")

if debug == 1: print("Apertura excel in corso!\n")

#rename
#print("Rinomino il file alla data odierna")
#currentDate = (date.today().strftime("%d-%b-%Y")).lower()
# Month abbreviation, day and year
#oldName = "D:\Documenti\Davide\GDriveMagistrale\TIM\\2021\Copertura attiva e pianificata FTTCab.xlsx"	
#newName = "D:\Documenti\Davide\GDriveMagistrale\TIM\\2021\Copertura attiva e pianificata FTTCab " + currentDate + ".xlsx"
#print("New file name: " + newName + "\n")
#os.rename(oldName, newName) 

#begin to open file and search
listOfFiles = glob.glob('D:\Documenti\Davide\GDriveMagistrale\TIM\\2021\*.xlsx')
lastFile = max(listOfFiles, key=os.path.getctime)
filePath = pathlib.Path(lastFile)
strPath = str(filePath)
print("Ho il percorso!\n" + strPath + "\nCaricamento File in corso! Attendere prego!\n")
file = load_workbook(strPath)
print("Caricamento foglio Excel in corso\n")
fttc = file['FTTC']
print("File caricato, inizio ricerca!")

#this function search for the specific cabinet CIBEITAC014, that is the one in my interests
for x in range(150000):
    if x > 0:
        if fttc.cell(row=x, column=5).value == "CIBEITAC014":
            if fttc.cell(row=x, column=9).value == "Pianificato":
                if fttc.cell(row=x, column=8).value == "100M":
                    print("Il cabinet NON è attivo. Data attivazione prevista: "+str(fttc.cell(row=x, column=11).value) + "\n")
                    print("Velocità di attivazione prevista: " + str(fttc.cell(row=x, column=8).value))
                    break
                elif fttc.cell(row=x, column=8).value == "Upgrade 200M":
                    print("Programmazione upgrade cabinet! Data prevista upgrade 200M: " + str(fttc.cell(row=x, column=11).value) + "\n")
                    break
            elif fttc.cell(row=x, column=9).value == "Attivo":
                if fttc.cell(row=x, column=8).value == "100M":
                    print("Il cabinet è attivo. Data attivazione prevista: "+str(fttc.cell(row=x, column=10).value) + "\n")
                    print("Velocità di attivazione prevista: " + str(fttc.cell(row=x, column=8).value))
                    break
                elif fttc.cell(row=x, column=8).value == "Upgrade 200M":
                    print("Programmazione upgrade cabinet! Data attivazione 200M: " + str(fttc.cell(row=x, column=10).value) + "\n")
                    break
            elif fttc.cell(row=x, column=9).value == "Sospeso":
                print("Il cabinet è SOSPESO, attendi\n")
                break
            elif fttc.cell(row=x, column=9).value == "Saturo":
                print("Cabinet SATURO, attendere desaturazione, probabile upgrade 200M\n")
                break
            time.sleep(10)
        else:
            x = x + 1
    else:
        x = 1

print("Fine del programma, alla prossima!")
